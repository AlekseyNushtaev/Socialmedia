import random
import os
import tempfile
from datetime import datetime, date, timezone, timedelta
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill

from bot import sql, x3, bot
from config import ADMIN_IDS, CHECKER_ID, API_FREEKASSA, SHOP_ID_FREEKASSA
from lexicon import lexicon
from keyboard import create_kb, STYLE_PRIMARY, STYLE_SUCCESS, STYLE_DANGER, keyboard_sub_after_buy, BTN_BACK
from logging_config import logger
import asyncio
from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton, FSInputFile
from aiogram.filters import Command, StateFilter

from handlers.handlers_broadcast import BroadcastState

from sheduler.check_connect import check_connect
from sheduler.check_fk import fetch_fk_payment_check_status
from payments.pay_freekassa import FreekassaPayment
from telegram_ids import is_telegram_chat_id
from X3 import panel_username_for_site_user
from wl_traffic.service import (
    fetch_panel_user,
    fetch_wl_traffic_gb_for_day,
    get_wl_used_gb_for_user,
    reassign_to_active_squad,
    user_on_active_squad,
    user_on_limited_squad,
)

router = Router()

_EXCEL_COL_WIDTH_MAX = 255


def _parse_check_fk_date(raw: str) -> Optional[date]:
    for fmt in ("%d.%m.%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(raw.strip(), fmt).date()
        except ValueError:
            continue
    return None


_ADD7ALL_PREVIEW_CB = "add7all_preview"
_ADD7ALL_YES_CB = "add7all_yes"
_ADD7ALL_NO_CB = "add7all_no"

_ADD7ALL_PROMO_TEXT = '''
💚<b>Самое время вернутся в Ускоритель СоцСетей — дарим 7 дней тестдрайва новых серверов</b>🔝
Подключение займет пару секунд

Жми👇
'''

_ADD7ALL_TRIAL_KB = create_kb(
    1,
    styles={"trial_return_get": STYLE_SUCCESS},
    trial_return_get="🔥Получить ТРИАЛ",
)

_ADD7SUB_TEXT = (
    "Уважаемые друзья!\n"
    "Мы столкнулись с аварией в датацентре.\n"
    "Проблема решена - бот и личный кабинет снова заработал.\n\n"
    "В качестве компенсации добавляем Вам 7 дней к подписке!\n\n"
    "Для обновления подписки:\n"
    '1. Нажмите кнопку "🔗 Подключить VPN"\n'
    '2. Перейдите в личный кабинет и нажмите кнопку "Добавить подписку"\n'
    "3. Подписка в Вашем приложении (Happ, др.) обновится"
)

_ADD7SUB_CONNECT_KB = create_kb(
    1,
    styles={"connect_vpn": STYLE_PRIMARY},
    connect_vpn="🔗 Подключить VPN",
)

_ADD7SUB_PROGRESS_EVERY = 1000
_ADD_TRAFFIC_ALL_PROGRESS_EVERY = 100

_ADD2BONUS_CUTOFF = datetime(2026, 6, 23, 0, 0, 0)
_ADD2BONUS_YES_CB = "add2bonus_yes"
_ADD2BONUS_NO_CB = "add2bonus_no"
_ADD2BONUS_PROGRESS_EVERY = 1000

_ADD2BONUS_TEXT = (
    "Дорогие друзья! 👋\n\n"
    "Мы столкнулись со <b>сбоем работы страницы подписки</b>.\n"
    "🔧 Технические работы уже ведутся — скоро всё устраним.\n\n"
    "В качестве компенсации дарим вам <b>+2 дня</b> к подписке! 🎁\n\n"
    "💡 <i>PS:</i> вы всегда можете импортировать настройки в приложение "
    "без перехода на страницу подписки — нажмите кнопку "
    "«⚠️ Если страница не загружается» и выполните пошаговую инструкцию для подключения."
)

_ADD2BONUS_KB = InlineKeyboardMarkup(
    inline_keyboard=[
        [
            InlineKeyboardButton(
                text="⚠️ Если страница не загружается",
                callback_data="import",
                style=STYLE_DANGER,
            )
        ]
    ]
)

_SUB_TIER_LABELS = {
    "main": "💫 Подписка PRO — соцсети",
    "white": "📱 Мобильный тариф",
}

_MSK = timezone(timedelta(hours=3))


def _msk_dt_str(dt: Optional[datetime]) -> str:
    if dt is None:
        return "Нет"
    if dt.tzinfo is None:
        aware = dt.replace(tzinfo=timezone.utc)
    else:
        aware = dt.astimezone(timezone.utc)
    return aware.astimezone(_MSK).strftime("%d-%m-%Y %H:%M МСК")


def _pay_dt_str(dt: Optional[datetime]) -> str:
    """Формат даты для /pay: YYYY-MM-DD HH:MM:SS (МСК)."""
    if dt is None:
        return "Нет"
    if dt.tzinfo is None:
        aware = dt.replace(tzinfo=timezone.utc)
    else:
        aware = dt.astimezone(timezone.utc)
    return aware.astimezone(_MSK).strftime("%Y-%m-%d %H:%M:%S")


def _pay_panel_sub_line(activ_result: dict) -> str:
    t = activ_result.get("time", "-")
    if t in (None, "", "-"):
        return "Нет"
    try:
        parsed = datetime.strptime(str(t).replace(" МСК", "").strip(), "%d-%m-%Y %H:%M")
        return parsed.strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return str(t)


def _panel_usernames_from_row(row: tuple) -> tuple[str, str]:
    """Пара username в панели: обычная, вайт (как в web_api._panel_vpn_usernames)."""
    tg_col = row[1]
    linked = row[28]
    tg = None
    if tg_col is not None and int(tg_col) > 0:
        tg = int(tg_col)
    elif linked is not None and int(linked) > 0:
        tg = int(linked)
    if tg is not None:
        s = str(tg)
        return s, f"{s}_white"
    db_uid = int(tg_col)
    return panel_username_for_site_user(db_uid, False), panel_username_for_site_user(db_uid, True)


def _split_long_text(text: str, limit: int = 3800) -> list[str]:
    if len(text) <= limit:
        return [text]
    parts: list[str] = []
    rest = text
    while rest:
        parts.append(rest[:limit])
        rest = rest[limit:]
    return parts


@router.message(F.video, F.from_user.id.in_(ADMIN_IDS), ~StateFilter(BroadcastState))
async def get_video(message: Message):
    await message.answer(message.video.file_id)


@router.message(F.photo, F.from_user.id.in_(ADMIN_IDS), ~StateFilter(BroadcastState))
async def get_photo(message: Message):
    await message.answer(message.photo[-1].file_id)


@router.message(Command(commands=['user']))
async def user_info(message: Message):

    # Проверка прав администратора
    if message.from_user.id not in ADMIN_IDS:
        return

    try:
        # Извлекаем аргументы команды
        args = message.text.split()

        if len(args) < 2:
            await message.answer("❌ Использование: /user <telegram_id>\nНапример: /user 123456789")
            return

        user_id = int(args[1].strip())

        # Проверяем, существует ли пользователь в БД
        user_data = await sql.get_user(user_id)

        if not user_data:
            await message.answer(f"❌ Пользователь с ID {user_id} не найден в базе данных.")
            return
        text = []
        for i in range(len(user_data)):
            if isinstance(user_data[i], datetime):
                item = user_data[i].strftime('%Y-%m-%d %H:%M:%S')
                text.append(item)
            elif user_data[i] is None:
                text.append('None')
            else:
                text.append(str(user_data[i]))
        text = '\n'.join(text)
        await message.answer(text)
    except Exception as e:
        await message.answer(f'Ошибка при формировании сообщения: {str(e)}')


@router.message(Command(commands=['gift']))
async def gift_info_command(message: Message):
    """Информация о подарке по gift_id."""
    if message.from_user.id not in ADMIN_IDS:
        return

    args = (message.text or "").split()
    if len(args) < 2:
        await message.answer("❌ Использование: /gift <gift_id>")
        return

    gift_id = args[1].strip()
    gift = await sql.get_gift(gift_id)
    if gift is None:
        await message.answer("Такого подарка нет.")
        return

    if gift.recepient_id is None:
        await message.answer("Подарок ещё не активирован.")
        return

    if gift.recepient_id > 0:
        await message.answer(f"Подарок активировал юзер с tg_id = {gift.recepient_id}.")
        return

    user = await sql.get_user_object_by_user_id(gift.recepient_id)
    panel_username = user.field_str_2 if user else None
    if panel_username:
        await message.answer(f"Подарок активировал юзер с username в панели {panel_username}")
    else:
        await message.answer(
            f"Подарок активирован {gift.recepient_id}"
        )


@router.message(Command(commands=['pay']))
async def pay_info_command(message: Message):
    """Сводка подписок (БД / панель) и успешные платежи пользователя."""
    if message.from_user.id not in ADMIN_IDS:
        return

    args = (message.text or "").split()
    if len(args) < 2:
        await message.answer("❌ Использование: /pay <telegram_id>\nНапример: /pay 123456789")
        return

    try:
        target_id = int(args[1].strip())
    except ValueError:
        await message.answer("❌ ID должен быть числом.")
        return

    user_row = await sql.get_user(target_id)
    if not user_row:
        await message.answer(f"❌ Пользователь {target_id} не найден в базе данных.")
        return

    reg_un, white_un = _panel_usernames_from_row(user_row)
    sub_db = user_row[9]
    white_db = user_row[10]

    try:
        ar_reg, ar_white = await asyncio.gather(
            x3.activ(reg_un),
            x3.activ(white_un),
        )
    except Exception as e:
        logger.exception("/pay: панель")
        await message.answer(f"❌ Ошибка запроса к панели: {e}")
        return

    pay_rows = await sql.get_user_subscription_payment_report(target_id)
    pay_lines: list[str] = []
    for tc, kind, method, detail in pay_rows:
        ts = _pay_dt_str(tc)
        pay_lines.append(f"• {ts} — {kind} — {method} — {detail}")

    trafic_wl, limit_wl = await sql.get_wl_limits(target_id)
    used_wl_gb = await get_wl_used_gb_for_user(x3, target_id, trafic_wl)
    remaining_wl = max(0.0, round(limit_wl - used_wl_gb, 2))

    body = (
        f"<b>/pay {target_id}</b>\n\n"
        f"Подписка обычная в БД бота — {_pay_dt_str(sub_db)}\n"
        f"Подписка обычная в панели — {_pay_panel_sub_line(ar_reg)}\n"
        f"Подписка вайт в БД бота — {_pay_dt_str(white_db)}\n"
        f"Подписка вайт в панели — {_pay_panel_sub_line(ar_white)}\n\n"
        f"📡 <b>Дополнительный сервер (WL-трафик)</b>\n"
        f"├ Лимит: <b>{limit_wl:.2f} GB</b>\n"
        f"├ Использовано: <b>{used_wl_gb:.2f} GB</b>\n"
        f"└ Осталось: <b>{remaining_wl:.2f} GB</b>\n\n"
        f"<b>Платежи:</b>\n"
    )
    if pay_lines:
        body += "\n".join(pay_lines)
    else:
        body += "Нет"

    for chunk in _split_long_text(body):
        await message.answer(chunk, parse_mode="HTML")


async def _partner_admin_stats_text(tg_id: int) -> Optional[str]:
    user = await sql.get_user_object_by_user_id(tg_id)
    if user is None:
        return None
    if not user.partner_flag:
        return "not_partner"

    referrals = await sql.select_partner_count(tg_id)
    payments_sum = await sql.select_partner_referrals_payments_sum(tg_id)
    balance = user.partner_balance or 0
    paid_out = user.partner_pay or 0
    total_earned = balance + paid_out

    return (
        f"📊 <b>Статистика {tg_id}:</b>\n\n"
        f"👥 Друзей перешло (/start): <b>{referrals}</b>\n"
        f"💳 Приобретено подписок друзьями на: <b>{payments_sum} ₽</b>\n\n"
        f"💵 Заработок партнёра (всего): <b>{total_earned} ₽</b>\n"
        f"✅ Выведено: <b>{paid_out} ₽</b>\n"
        f"🏦 Осталось на вывод: <b>{balance} ₽</b>"
    )


@router.message(Command(commands=['partner']))
async def partner_info_command(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        return

    args = (message.text or "").split()
    if len(args) < 2:
        await message.answer(
            "❌ Использование: /partner <telegram_id>\nНапример: /partner 123456789"
        )
        return

    try:
        target_id = int(args[1].strip())
    except ValueError:
        await message.answer("❌ ID должен быть числом.")
        return

    try:
        text = await _partner_admin_stats_text(target_id)
    except Exception as e:
        logger.exception("/partner")
        await message.answer(f"❌ Ошибка: {e}")
        return

    if text is None:
        await message.answer(f"❌ Пользователь {target_id} не найден в базе данных.")
        return
    if text == "not_partner":
        await message.answer(
            f"❌ Пользователь {target_id} не участвует в партнёрской программе "
            f"(partner_flag = False)."
        )
        return

    await message.answer(text, parse_mode="HTML")


@router.message(Command(commands=['partner_remove']))
async def partner_remove_command(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        return

    args = (message.text or "").split()
    if len(args) < 3:
        await message.answer(
            "❌ Использование: /partner_remove <telegram_id> <сумма>\n"
            "Например: /partner_remove 123456789 500"
        )
        return

    try:
        target_id = int(args[1].strip())
        amount = int(args[2].strip())
    except ValueError:
        await message.answer("❌ ID и сумма должны быть целыми числами.")
        return

    ok, err = await sql.partner_record_payout(target_id, amount)
    if not ok:
        await message.answer(f"❌ {err}")
        return

    stats = await _partner_admin_stats_text(target_id)
    if stats and stats != "not_partner":
        await message.answer(
            f"✅ Списано <b>{amount} ₽</b> с баланса, добавлено в «Выведено».\n\n{stats}",
            parse_mode="HTML",
        )
    else:
        await message.answer(
            f"✅ Списано {amount} ₽ с баланса пользователя {target_id}, добавлено в partner_pay."
        )


@router.message(Command(commands=['sub']))
async def set_subscription_date(message: Message):
    """Установка subscription_end_date или white_subscription_end_date в БД и панели"""
    if message.from_user.id not in ADMIN_IDS:
        await message.answer("❌ Эта команда доступна только администраторам.")
        return

    try:
        args = message.text.split()
        if len(args) < 3:
            await message.answer(
                "❌ Использование:\n"
                "  /sub <telegram_id> <дата_время>               – обновить обычную подписку\n"
                "  /sub <telegram_id> white <дата_время>         – обновить белую подписку\n"
                "Примеры:\n"
                "  /sub 123456789 2026-02-01 17:14:27\n"
                "  /sub 123456789 white 2026-02-01 17:14:27\n"
                "Формат даты: YYYY-MM-DD HH:MM:SS"
            )
            return

        user_id = int(args[1].strip())

        # Определяем тип и позицию даты
        if args[2].lower() == 'white':
            is_white = True
            date_str = " ".join(args[3:])
        else:
            is_white = False
            date_str = " ".join(args[2:])

        # Парсим дату
        date_formats = [
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%d.%m.%Y %H:%M:%S",
            "%d.%m.%Y %H:%M"
        ]
        target_date = None
        for fmt in date_formats:
            try:
                target_date = datetime.strptime(date_str, fmt)
                target_date = target_date.replace(tzinfo=timezone.utc)  # панель работает в UTC
                break
            except ValueError:
                continue
        if target_date is None:
            await message.answer(f"❌ Неверный формат даты: {date_str}")
            return

        # Проверяем наличие пользователя в БД
        user_data = await sql.get_user(user_id)
        if not user_data:
            await message.answer("⚠️ Пользователь не найден в БД.")
            return

        # Формируем username для панели
        username = str(user_id) + ('_white' if is_white else '')

        # Устанавливаем дату в панели
        success, actual_date = await x3.set_expiration_date(username, target_date, user_id)

        if not success or actual_date is None:
            await message.answer("❌ Не удалось установить дату в панели. Подробности в логах.")
            return

        if is_white:
            await sql.update_white_subscription_end_date(user_id, actual_date)
        else:
            await sql.update_subscription_end_date(user_id, actual_date)

        tier = "white" if is_white else "main"
        notify_status = ""
        if is_telegram_chat_id(user_id):
            try:
                sub_link = await x3.sublink(username)
                user_text = lexicon["sub_granted_notify"].format(
                    tier=_SUB_TIER_LABELS.get(tier, tier),
                    end_date=_msk_dt_str(actual_date),
                )
                await bot.send_message(
                    user_id,
                    user_text,
                    parse_mode="HTML",
                    disable_web_page_preview=True,
                    reply_markup=keyboard_sub_after_buy(sub_link) if sub_link else None,
                )
                notify_status = "\n📨 Пользователь уведомлён."
            except Exception as e:
                logger.error(f"/sub: не удалось уведомить user={user_id}: {e}")
                notify_status = f"\n⚠️ Не удалось уведомить пользователя: {e}"
        else:
            notify_status = "\nℹ️ Уведомление не отправлено (не Telegram ID)."

        await message.answer(
            f"✅ Дата подписки успешно установлена!\n\n"
            f"👤 Пользователь: {user_id}\n"
            f"📅 Целевая дата (UTC): {target_date.strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"📅 Установленная в панели дата (UTC): {actual_date.strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"📝 Тип: {_SUB_TIER_LABELS.get(tier, tier)}\n"
            f"💾 База данных обновлена."
            f"{notify_status}"
        )

    except Exception as e:
        logger.error(f"Ошибка в команде /sub: {e}")
        await message.answer(f"❌ Произошла ошибка: {str(e)}")


@router.message(Command(commands=['delete']))
async def delete_user_command(message: Message):
    """Удаление пользователя из БД по Telegram ID"""

    # Проверка прав администратора
    if message.from_user.id not in ADMIN_IDS:
        return

    try:
        # Извлекаем аргументы команды
        args = message.text.split()

        if len(args) < 2:
            await message.answer("❌ Использование: /delete <telegram_id>\nНапример: /delete 123456789")
            return

        user_id_to_delete = int(args[1].strip())

        # Проверяем, существует ли пользователь в БД
        user_data = await sql.get_user(user_id_to_delete)

        if not user_data:
            await message.answer(f"❌ Пользователь с ID {user_id_to_delete} не найден в базе данных.")
            return

        # Получаем информацию о пользователе для уведомления
        user_info = {
            "user_id": user_data[1],  # user_id
            "ref": user_data[2],  # ref
            "in_panel": user_data[4],  # in_panel
            "in_chanel": user_data[7] if len(user_data) > 7 else False  # in_chanel
        }

        # УДАЛЯЕМ ПОЛЬЗОВАТЕЛЯ ИЗ БД
        deletion_success = await sql.delete_from_db(user_id_to_delete)

        if deletion_success:
            # Логируем действие
            logger.info(f"Администратор {message.from_user.id} удалил пользователя {user_id_to_delete} из БД")

            # Формируем отчет об удалении
            report_message = (
                f"✅ Пользователь успешно удалён из базы данных\n\n"
                f"📋 Информация об удалённом пользователе:\n"
                f"├ ID: {user_info['user_id']}\n"
                f"├ Реферер: {user_info['ref'] if user_info['ref'] else 'нет'}\n"
                f"└ Брал ключ: {'✅ да' if user_info['in_panel'] else '❌ нет'}\n"
                f"⚠️ Пользователь удалён только из базы данных бота.\n"
                f"   Подписка в панели управления (X3) остаётся активной.\n"
                f"   Чтобы удалить полностью, используйте команду /gift на 0 дней."
            )

            await message.answer(report_message)

        else:
            await message.answer(f"❌ Ошибка при удалении пользователя {user_id_to_delete}.\n"
                                 "Возможно, пользователь уже был удалён или произошла ошибка базы данных.")

    except ValueError:
        await message.answer("❌ Неверный формат Telegram ID.\n"
                             "Используйте только цифры, например: /delete 123456789")
    except Exception as e:
        logger.error(f"Ошибка в команде /delete: {e}")
        await message.answer(f"❌ Произошла ошибка при выполнении команды: {str(e)}")


@router.message(Command("online"))
async def check_online(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        return
    users_x3 = await x3.get_all_users()

    active_telegram_ids = []
    for user in users_x3:
        if user['userTraffic']['firstConnectedAt']:
            connected_str = user['userTraffic']['onlineAt']
            try:
                connected_dt = datetime.fromisoformat(connected_str.replace('Z', '+00:00'))
                connected_date = connected_dt.date()
                if connected_date == datetime.now().date():
                    telegram_id = user.get('telegramId')
                    if telegram_id is not None:
                        active_telegram_ids.append(int(telegram_id))
            except (ValueError, TypeError):
                continue

    count_pay = 0
    count_trial = 0
    for tg_id in active_telegram_ids:
        user_data = await sql.get_user(tg_id)
        if user_data:
            if user_data[8]:
                count_pay += 1
            else:
                count_trial += 1
    await message.answer(
        f"Всего юзеров в панели: {len(users_x3)}\n"
        f"Юзеров, которые были онлайн сегодня: {len(active_telegram_ids)}\n"
        f"Юзеры с платной подпиской: {count_pay}\n"
        f"Юзеры на триале: {count_trial}"
    )


@router.message(Command("balance_panel"))
async def check_online(message: Message):
    squad_1 = ['28b6a3bf-8e81-42dd-9ac8-dab1c9a60b0a']
    squad_2 = ['85f8520a-8dd7-40a6-9f27-8a7467096c6a']
    success_count = 0
    fail_count = 0
    if message.from_user.id not in ADMIN_IDS:
        return
    users_x3 = await x3.get_all_users()
    for user in users_x3:
        await asyncio.sleep(0.3)
        random_squad = random.choice([squad_1, squad_2])
        username = user.get('username', '')
        if 'white' not in username and 'cascade-bridge-system' not in username:
            panel_user_id = user.get('id')
            connect = user.get('firstConnectedAt')
            if panel_user_id is not None and connect:
                if await x3.update_user_squads(int(panel_user_id), random_squad):
                    success_count += 1
                else:
                    fail_count += 1
    await message.answer(f"{len(users_x3)} - всего юзеров в панели\n{success_count + fail_count} - подключенных\n{success_count} - обновлено\n{fail_count} - ошибка")


@router.message(Command(commands=['sync_panel']))
async def sync_panel(message: Message):
    if message.from_user.id not in ADMIN_IDS:
        return

    await message.answer("🔄 Запускаю синхронизацию пользователей...")

    # 1. Получаем всех пользователей из панели и строим словарь {telegramId: user_data}
    users_panel = await x3.get_all_users()
    panel_dict = {}
    for user in users_panel:
        tg_id = user.get('telegramId')
        if tg_id is not None:
            panel_dict[tg_id] = user

    # 2. Получаем список пользователей, у которых is_pay_null=True и subscription_end_date=None
    users_for_sync = await sql.select_subscribed_not_in_chanel()

    # 3. Статистика
    updated = 0          # обновлено дат в БД
    added_to_panel = 0   # добавлено в панель
    not_found = 0        # не найдено в панели (остались в списке)

    # 4. Обрабатываем каждого пользователя из списка на синхронизацию
    if CHECKER_ID is not None:
        await bot.send_message(CHECKER_ID,
                               'Добрый день. Мы создали Вам личный кабинет и начислили 5 дней пробного '
                               'доступа.\nПерейдите по ссылке, нажав на кнопку 🌐 Подключить Ускоритель соцсетей',
                               reply_markup=create_kb(
                                   1,
                                   styles={'connect_vpn': STYLE_PRIMARY},
                                   connect_vpn='🌐 Подключить Ускоритель соцсетей',
                               ))

    for user_id in users_for_sync:
        # Проверяем, есть ли пользователь в панели
        if user_id in panel_dict:
            user_data = panel_dict[user_id]

            # Получаем expireAt и преобразуем в datetime
            expire_str = user_data.get('expireAt')
            if expire_str:
                try:
                    expire_dt = datetime.fromisoformat(expire_str.replace('Z', '+00:00'))
                except Exception as e:
                    logger.error(f"Ошибка парсинга expireAt для {user_id}: {e}")
                    continue

                await sql.update_subscription_end_date(user_id, expire_dt)
                updated += 1
                logger.info(f"Обновлена дата для {user_id} до {expire_dt}")
        else:
            user_id_str = str(user_id)
            result = await x3.addClient(5, user_id_str, user_id)
            if result:
                added_to_panel += 1
                logger.info(f"Добавлен в панель пользователь {user_id} (day=0)")
                await bot.send_message(user_id,
                                       'Добрый день. Мы создали Вам личный кабинет и начислили 5 дней пробного '
                                       'доступа.\nПерейдите по ссылке, нажав на кнопку 🌐 Подключить Ускоритель соцсетей',
                                       reply_markup=create_kb(
                                           1,
                                           styles={'connect_vpn': STYLE_PRIMARY},
                                           connect_vpn='🌐 Подключить Ускоритель соцсетей',
                                       ))
            else:
                not_found += 1
                logger.warning(f"Не удалось добавить в панель пользователя {user_id}")

    # 5. Итоговый отчёт
    report = (
        f"✅ Синхронизация завершена.\n"
        f"📊 Всего в панели: {len(users_panel)}\n"
        f"📋 Ожидало синхронизации: {len(users_for_sync)}\n"
        f"🔄 Обновлено дат в БД: {updated}\n"
        f"➕ Добавлено в панель (day=5): {added_to_panel}\n"
        f"❌ Не удалось добавить (ошибки): {not_found}"
    )
    await message.answer(report)
    logger.info(report)


@router.message(Command(commands=['shortuuid_export']))
async def shortuuid_export(message: Message):
    """Синхронизация shortUuid из панели в поля subscribtion / white_subscription в БД."""
    if message.from_user.id not in ADMIN_IDS:
        return

    await message.answer("🔄 Загружаю пользователей панели и записываю shortUuid в БД...")

    try:
        panel_users = await x3.get_all_users()
    except Exception as e:
        logger.error(f"shortuuid_export: панель: {e}")
        await message.answer(f"❌ Ошибка при запросе панели: {e}")
        return

    updated_sub = 0
    updated_white = 0
    skip_no_db = 0
    skip_no_tg = 0
    skip_no_short = 0
    errors = 0

    for user in panel_users:
        tg_id = user.get("telegramId")
        username = user.get("username") or ""
        if tg_id is None:
            if username.isdigit():
                tg_id = int(username)
            else:
                skip_no_tg += 1
                continue
        else:
            tg_id = int(tg_id)

        short_uuid = user.get("shortUuid")
        if not short_uuid:
            skip_no_short += 1
            continue

        db_user = await sql.get_user(tg_id)
        if not db_user:
            skip_no_db += 1
            continue

        is_white = "white" in username
        try:
            if is_white:
                await sql.update_white_subscription(tg_id, short_uuid)
                updated_white += 1
            else:
                await sql.update_subscribtion(tg_id, short_uuid)
                updated_sub += 1
            logger.success(f"shortuuid_export user {tg_id}: {short_uuid}")
        except Exception as e:
            errors += 1
            logger.error(f"shortuuid_export user {tg_id}: {e}")

    report = (
        f"✅ Готово.\n"
        f"📊 В панели записей: {len(panel_users)}\n"
        f"📝 subscribtion обновлено: {updated_sub}\n"
        f"📝 white_subscription обновлено: {updated_white}\n"
        f"⏭ без telegramId/username: {skip_no_tg}\n"
        f"⏭ без shortUuid: {skip_no_short}\n"
        f"⏭ нет в БД: {skip_no_db}\n"
        f"❌ ошибок записи: {errors}"
    )
    await message.answer(report)
    logger.info(report)


@router.message(Command(commands=['check_users']))
async def check_users_command(message: Message):
    """Проверка соответствия дат окончания подписки у оплаченных пользователей (has_discount=True)"""
    if message.from_user.id not in ADMIN_IDS:
        return

    await message.answer("🔄 Начинаю проверку пользователей с оплатами...")

    try:
        # 1. Получаем список оплаченных пользователей из БД
        users_with_discount = await sql.get_users_with_payment()
        total = len(users_with_discount)
        if total == 0:
            await message.answer("❌ Нет пользователей с оплатами.")
            return

        # 2. Получаем всех пользователей из панели (один запрос)
        panel_users = await x3.get_all_users()
        logger.info(f"Загружено {len(panel_users)} пользователей из панели")

        # 3. Строим словарь для быстрого поиска по telegramId и username
        panel_by_telegram = {}      # ключ: telegramId (int)
        panel_by_username = {}      # ключ: username (str)

        for user in panel_users:
            tg_id = user.get('telegramId')
            username = user.get('username')
            if tg_id is not None:
                panel_by_telegram[int(tg_id)] = user
            elif username:
                panel_by_username[username] = user

        # 4. Проходим по всем оплаченным пользователям и ищем их в панели
        mismatched = []      # кортежи (user_id, db_date, panel_date) для расхождений >=3ч
        not_found_in_panel = []  # пользователи, отсутствующие в панели
        processed = 0

        for user_id in users_with_discount:
            processed += 1
            if processed % 10 == 0:
                logger.info(f"Проверено {processed}/{total}")

            # Пытаемся найти пользователя в панели
            panel_user = panel_by_telegram.get(user_id)
            if panel_user is None:
                panel_user = panel_by_username.get(str(user_id))

            if panel_user is None:
                not_found_in_panel.append(user_id)
                continue

            expire_str = panel_user.get('expireAt')
            if not expire_str:
                # нет даты в панели – считаем расхождением (panel_date = None)
                db_expire = await sql.get_subscription_end_date(user_id)
                mismatched.append((user_id, db_expire, None))
                continue

            try:
                panel_expire = datetime.fromisoformat(expire_str.replace('Z', '+00:00'))
            except Exception:
                # не удалось распарсить дату панели
                db_expire = await sql.get_subscription_end_date(user_id)
                mismatched.append((user_id, db_expire, None))
                continue

            # Получаем дату из БД (обычная подписка)
            db_expire = await sql.get_subscription_end_date(user_id)
            panel_naive = panel_expire.replace(tzinfo=None)

            if db_expire is None:
                # нет даты в БД
                mismatched.append((user_id, None, panel_naive))
                continue

            db_naive = db_expire.replace(tzinfo=None)
            diff_hours = abs((panel_naive - db_naive).total_seconds()) / 3600

            if diff_hours >= 6:
                mismatched.append((user_id, db_naive, panel_naive))

        # 5. Формируем отчёт
        report_lines = []
        report_lines.append(f"📊 Результаты проверки:\n")
        report_lines.append(f"👥 Всего проверено: {total}")
        report_lines.append(f"❌ Расхождений в датах (>=6ч): {len(mismatched)}")
        report_lines.append(f"🔍 Не найдены в панели: {len(not_found_in_panel)}")

        # Если есть расхождения и их количество не превышает лимит для прямого вывода
        if mismatched or not_found_in_panel:
            if len(mismatched) <= 50 and len(not_found_in_panel) <= 50:
                if mismatched:
                    report_lines.append("\n🆔 Расхождения (команды для синхронизации):")
                    for uid, db_dt, panel_dt in mismatched:
                        db_str = db_dt.strftime('%Y-%m-%d %H:%M:%S') if db_dt else 'None'
                        panel_str = panel_dt.strftime('%Y-%m-%d %H:%M:%S') if panel_dt else 'None'
                        report_lines.append(f"/sub {uid} {db_str} /sub {uid} {panel_str}")
                if not_found_in_panel:
                    report_lines.append("\n🆔 Не найдены в панели:")
                    report_lines.extend(str(uid) for uid in not_found_in_panel)
                await message.answer("\n".join(report_lines))
            else:
                # Если много расхождений – отправляем файлом
                import io
                text_io = io.StringIO()
                text_io.write("user_id\tdb_date\tpanel_date\n")
                for uid, db_dt, panel_dt in mismatched:
                    db_str = db_dt.strftime('%Y-%m-%d %H:%M:%S') if db_dt else 'None'
                    panel_str = panel_dt.strftime('%Y-%m-%d %H:%M:%S') if panel_dt else 'None'
                    text_io.write(f"/sub {uid} {db_str} /sub {uid} {panel_str}\n")
                for uid in not_found_in_panel:
                    text_io.write(f"{uid}\tnot_found\n")
                text_io.seek(0)
                from aiogram.types import BufferedInputFile
                file_data = BufferedInputFile(text_io.getvalue().encode(), filename="check_users_report.txt")
                await message.answer_document(
                    document=file_data,
                    caption="\n".join(report_lines[:5])
                )
        else:
            await message.answer("✅ Все оплаченные пользователи синхронизированы (разница менее 3 часов).")

    except Exception as e:
        logger.exception("Ошибка в /check_users")
        await message.answer(f"❌ Ошибка: {str(e)}")


@router.message(Command(commands=['check_fk']))
async def check_fk_command(message: Message):
    """Проверка платежей FreeKassa за указанный день через API; выгрузка в Excel."""
    if message.from_user.id not in ADMIN_IDS:
        return

    args = message.text.split(maxsplit=1)
    if len(args) < 2:
        await message.answer(
            "❌ Использование: /check_fk DD.MM.YY\n"
            "Пример: /check_fk 01.08.25"
        )
        return

    day = _parse_check_fk_date(args[1])
    if day is None:
        await message.answer(f"❌ Неверный формат даты: {args[1]}\nОжидается DD.MM.YY или DD.MM.YYYY")
        return

    if not API_FREEKASSA or SHOP_ID_FREEKASSA is None:
        await message.answer("❌ FreeKassa API не настроен (API_FREEKASSA / SHOP_ID_FREEKASSA).")
        return

    await message.answer(
        f"🔄 Проверяю платежи FreeKassa за {day.strftime('%d.%m.%Y')} через API..."
    )

    export_path = None
    try:
        payments = await sql.get_fk_sbp_payments_for_date(day)
        if not payments:
            await message.answer(f"ℹ️ Платежей FreeKassa за {day.strftime('%d.%m.%Y')} не найдено.")
            return

        fk = FreekassaPayment(API_FREEKASSA, SHOP_ID_FREEKASSA)
        checked: list[tuple] = []
        for i, pay in enumerate(payments, 1):
            status_check = await fetch_fk_payment_check_status(pay, fk)
            checked.append((pay, status_check))
            if i % 10 == 0:
                await message.answer(f"⏳ Проверено {i}/{len(payments)}...")

        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        light_red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        fk_columns = [
            "ID", "User ID", "Amount", "Time Created", "Is Gift", "Status",
            "Status_check", "Transaction_Id", "FK_Order_Id", "Nonce", "Signature", "Method", "Payload",
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "payments_fk_sbp"
        for col_num, title in enumerate(fk_columns, 1):
            cell = ws.cell(row=1, column=col_num, value=title)
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_num, (pay, status_check) in enumerate(checked, 2):
            row_data = [
                pay.id, pay.user_id, pay.amount, pay.time_created,
                pay.is_gift, pay.status, status_check, pay.transaction_id, pay.fk_order_id,
                pay.nonce, pay.signature, pay.method, pay.payload,
            ]
            mismatch_row = pay.status != status_check
            for col_num, value in enumerate(row_data, 1):
                if col_num == 4 and value and isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
                if mismatch_row:
                    cell.fill = light_red_fill

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, _EXCEL_COL_WIDTH_MAX)

        export_path = tempfile.mktemp(suffix='.xlsx')
        wb.save(export_path)

        mismatch = sum(1 for pay, sc in checked if pay.status != sc)
        confirmed_api = sum(1 for _, sc in checked if sc == "confirmed")
        caption = (
            f"📊 Проверка FreeKassa за {day.strftime('%d.%m.%Y')}\n"
            f"📅 Создано: {datetime.now().strftime('%d.%m.%Y %H:%M')}\n\n"
            f"Всего платежей: {len(checked)}\n"
            f"Подтверждено по API: {confirmed_api}\n"
            f"Расхождение Status ≠ Status_check: {mismatch}"
        )
        await message.answer_document(
            document=FSInputFile(export_path, filename=f"check_fk_{day.strftime('%d.%m.%y')}.xlsx"),
            caption=caption,
        )
        logger.info(
            f"Админ {message.from_user.id} выполнил /check_fk {args[1]}: "
            f"{len(checked)} платежей, расхождений {mismatch}"
        )

    except Exception as e:
        logger.exception("Ошибка в /check_fk")
        await message.answer(f"❌ Ошибка: {str(e)}")
    finally:
        if export_path:
            try:
                os.remove(export_path)
            except OSError:
                pass


@router.message(Command(commands=['send_gift']))
async def send_gift_command(message: Message):
    """Отправляет подарок (3 дня подписки) пользователям, созданным 16 или 17 марта 2026,
    у которых in_panel=True, is_connect=False, is_delete=False."""
    if CHECKER_ID is None or message.from_user.id != CHECKER_ID:
        return

    await message.answer("🔄 Начинаю отправку подарков...")

    # Целевые даты
    target_dates = (datetime(2026, 3, 16), datetime(2026, 3, 17))

    # Получаем всех пользователей из БД (можно фильтровать на стороне Python, т.к. запрос сложный)
    all_users = await sql.get_all_users()  # список объектов Users

    # Фильтруем вручную
    candidates = [CHECKER_ID]
    for user in all_users:
        if user.is_delete:
            continue
        if not user.in_panel:
            continue
        if user.is_connect:
            continue
        if user.create_user.date() not in [d.date() for d in target_dates]:
            continue
        candidates.append(user.user_id)

    if not candidates:
        await message.answer("❌ Нет пользователей, удовлетворяющих условиям.")
        return
    else:
        await message.answer(f"Всего {len(candidates)} пользователей, удовлетворяющих условиям.")

    success_count = 0
    fail_count = 0

    # Текст сообщения
    gift_text = '''
🥵 Это была DDoS-атака!

Друзья, простите за временные неудобства. Сервис работает в штатном режиме.

Мы столкнулись с мощной DDoS-атакой, если у вас <b>не открывался личный кабинет — проблема уже решена.</b>

🔥 Мы начислили вам <b>дополнительные 5 дней</b> к подписке, чтобы вы могли оценить удобство Ускорителя соцсетей.

📱 Не можете настроить?
Если вы никак не могли разобраться с импортом конфигов — <b>смотрите видеоинструкцию</b>! Там всё разложено по полочкам.

🌐 Осталось только нажать кнопку "🌐 Подключить Ускоритель соцсетей" — и вы в деле.
            '''

    for user_id in candidates[83:]:
        try:
            # Отправляем сообщение
            await bot.send_message(user_id,
                                   gift_text,
                                   reply_markup=create_kb(
                                       1,
                                       styles={
                                           'video_faq': STYLE_PRIMARY,
                                           'connect_vpn': STYLE_PRIMARY,
                                       },
                                       video_faq='🎥 Видеоинструкция',
                                       connect_vpn='🌐 Подключить Ускоритель соцсетей',
                                   ))
            # Добавляем 3 дня подписки
            result = await x3.updateClient(5, str(user_id), user_id)
            if result:
                success_count += 1
                logger.info(f"Подарок отправлен пользователю {user_id}")
            else:
                fail_count += 1
                logger.error(f"Не удалось обновить подписку для {user_id}")
            await asyncio.sleep(0.05)  # небольшая задержка
        except Exception as e:
            fail_count += 1
            logger.error(f"Ошибка при обработке {user_id}: {e}")

    await message.answer(
        f"✅ Рассылка подарков завершена.\n"
        f"👥 Найдено: {len(candidates)}\n"
        f"✅ Успешно: {success_count}\n"
        f"❌ Ошибок: {fail_count}"
    )


_ADD_NEW_USERS_CUTOFF = datetime(2026, 6, 5, 0, 0, 0)
_ADD_NEW_USERS_PHASE2_EXPIRE = _ADD_NEW_USERS_CUTOFF
_ADD_NEW_USERS_PROGRESS_EVERY = 1000


async def _add_new_users_process_phase(
    users: list,
    expire_resolver,
    admin_chat_id: int,
    phase_label: str,
) -> dict:
    """Добавляет пользователей в панель; expire_resolver(user) -> datetime."""
    stats = {
        "total": len(users),
        "ok": 0,
        "fail": 0,
        "skipped_panel": 0,
        "skipped_non_tg": 0,
    }
    for idx, user in enumerate(users, start=1):
        uid = int(user.user_id)
        if not is_telegram_chat_id(uid):
            stats["skipped_non_tg"] += 1
            await asyncio.sleep(0.02)
            continue
        user_id_str = str(uid)
        panel_resp = await x3.get_user_by_username(user_id_str)
        if panel_resp and panel_resp.get("response"):
            stats["skipped_panel"] += 1
            await sql.update_in_panel(uid)
            await asyncio.sleep(0.02)
            continue
        expire_at = expire_resolver(user)
        short_uuid = (user.subscribtion or "").strip() or None
        ok = await x3.add_client_migrate(uid, expire_at, short_uuid=short_uuid)
        if ok:
            stats["ok"] += 1
        else:
            stats["fail"] += 1
        if idx % _ADD_NEW_USERS_PROGRESS_EVERY == 0:
            try:
                await bot.send_message(
                    admin_chat_id,
                    f"{phase_label}: обработано {idx}/{stats['total']}, "
                    f"добавлено {stats['ok']}, ошибок {stats['fail']}…",
                )
            except Exception as notify_err:
                logger.warning("%s: прогресс админу: %s", phase_label, notify_err)
        await asyncio.sleep(0.05)
    return stats


@router.message(Command(commands=["add_new_users"]))
async def add_new_users_command(message: Message):
    """
    Миграция в панель из БД:
    1) subscription_end_date > 2026-06-05 — expireAt из БД;
    2) остальные с датой — expireAt 2026-06-05 00:00:00.
    """
    if message.from_user.id not in ADMIN_IDS:
        return

    phase1 = await sql.select_users_subscription_after_cutoff(_ADD_NEW_USERS_CUTOFF)
    phase2 = await sql.select_users_subscription_on_or_before_cutoff(_ADD_NEW_USERS_CUTOFF)

    await message.answer(
        "⏳ /add_new_users\n"
        f"Фаза 1 (дата окончания > {_ADD_NEW_USERS_CUTOFF:%Y-%m-%d %H:%M:%S}): {len(phase1)} чел.\n"
        f"Фаза 2 (дата есть, ≤ порога): {len(phase2)} чел.\n"
        "Начинаю…"
    )
    admin_chat_id = message.chat.id

    def _expire_from_db(u):
        dt = u.subscription_end_date
        if dt.tzinfo is None:
            return dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)

    s1 = await _add_new_users_process_phase(
        phase1,
        _expire_from_db,
        admin_chat_id,
        "add_new_users фаза 1",
    )
    s2 = await _add_new_users_process_phase(
        phase2,
        lambda _u: _ADD_NEW_USERS_PHASE2_EXPIRE.replace(tzinfo=timezone.utc),
        admin_chat_id,
        "add_new_users фаза 2",
    )

    report = (
        "✅ /add_new_users завершено.\n\n"
        f"Фаза 1 (> {_ADD_NEW_USERS_CUTOFF:%Y-%m-%d}):\n"
        f"• В выборке: {s1['total']}\n"
        f"• Добавлено: {s1['ok']}\n"
        f"• Уже в панели: {s1['skipped_panel']}\n"
        f"• Ошибок: {s1['fail']}\n"
        f"• Не Telegram ID: {s1['skipped_non_tg']}\n\n"
        f"Фаза 2 (остальные с датой, expireAt {_ADD_NEW_USERS_PHASE2_EXPIRE:%Y-%m-%d}):\n"
        f"• В выборке: {s2['total']}\n"
        f"• Добавлено: {s2['ok']}\n"
        f"• Уже в панели: {s2['skipped_panel']}\n"
        f"• Ошибок: {s2['fail']}\n"
        f"• Не Telegram ID: {s2['skipped_non_tg']}"
    )
    await message.answer(report)
    logger.info(f"Админ {message.from_user.id} /add_new_users: {report}")


@router.message(Command(commands=['reset_field_bool_2']))
async def reset_field_bool_2_command(message: Message):
    """Сброс field_bool_2: у всех или у одного user_id."""
    if message.from_user.id not in ADMIN_IDS:
        return

    args = (message.text or "").split()
    if len(args) >= 2:
        try:
            target_id = int(args[1].strip())
        except ValueError:
            await message.answer("❌ Использование: /reset_field_bool_2 [telegram_id]")
            return
        user_row = await sql.get_user(target_id)
        if not user_row:
            await message.answer(f"❌ Пользователь {target_id} не найден.")
            return
        await sql.update_field_bool_2(target_id, False)
        await message.answer(f"Готово: field_bool_2 = false для user_id {target_id}.")
        logger.info(f"Админ {message.from_user.id}: сброс field_bool_2 для {target_id}")
        return

    n = await sql.reset_field_bool_2_all()
    await message.answer(f"Готово: field_bool_2 = false у {n} записей в users.")
    logger.info(f"Админ {message.from_user.id}: сброс field_bool_2 для всех, обновлено строк: {n}")


@router.message(Command(commands=['add_traffic']))
async def add_traffic_command(message: Message):
    """Админ: добавить GB к limit_wl, при необходимости вернуть на active squad."""
    if message.from_user.id not in ADMIN_IDS:
        return

    args = (message.text or "").split()
    if len(args) < 3:
        await message.answer(
            "❌ Использование: /add_traffic <telegram_id> <GB>\n"
            "Например: /add_traffic 123456789 10"
        )
        return

    try:
        target_id = int(args[1].strip())
        gb = float(args[2].strip().replace(",", "."))
    except ValueError:
        await message.answer("❌ ID и количество GB должны быть числами.")
        return

    if gb <= 0:
        await message.answer("❌ Количество GB должно быть больше 0.")
        return

    user_row = await sql.get_user(target_id)
    if not user_row:
        await message.answer(f"❌ Пользователь {target_id} не найден в базе данных.")
        return

    trafic_wl, _ = await sql.get_wl_limits(target_id)
    used_gb = await get_wl_used_gb_for_user(x3, target_id, trafic_wl)

    await sql.add_wl_limit(target_id, gb)
    _, limit_wl = await sql.get_wl_limits(target_id)
    remaining_gb = max(0.0, round(limit_wl - used_gb, 2))

    panel_user = await fetch_panel_user(x3, target_id)
    squad_note = ""
    if panel_user:
        user_row_after = await sql.get_user(target_id)
        field_bool_2 = bool(user_row_after[25]) if user_row_after else False
        if (
            user_on_limited_squad(panel_user)
            and limit_wl > used_gb
            and not field_bool_2
        ):
            if await reassign_to_active_squad(x3, panel_user):
                squad_note = "\n✅ Squad → active (Дополнительный сервер)"
            else:
                squad_note = "\n⚠️ Не удалось переназначить squad в панели"

    admin_text = (
        f"✅ <b>Добавлено {gb:g} GB</b> для user <code>{target_id}</code>{squad_note}\n\n"
        f"├ Использовано: <b>{used_gb:.2f} GB</b>\n"
        f"└ Лимит: <b>{limit_wl:.2f} GB</b>"
    )
    await message.answer(admin_text, parse_mode="HTML")
    logger.info(
        f"Админ {message.from_user.id}: /add_traffic uid={target_id} +{gb:g} GB "
        f"used={used_gb:.2f} limit={limit_wl:.2f}"
    )

    if target_id > 0:
        try:
            await bot.send_message(
                chat_id=target_id,
                text=lexicon["wl_traffic_admin_grant"].format(
                    gb=gb,
                    limit_gb=limit_wl,
                    used_gb=used_gb,
                    remaining_gb=remaining_gb,
                ),
                parse_mode="HTML",
                reply_markup=create_kb(1, back_to_main=BTN_BACK),
            )
        except Exception as e:
            await message.answer(f"⚠️ Лимит добавлен, но push пользователю не отправлен: {e}")
            logger.error(f"/add_traffic: push uid={target_id}: {e}")


@router.message(Command(commands=['add_traffic_all']))
async def add_traffic_all_command(message: Message):
    """+10 GB limit_wl всем с подпиской до конца сегодня или позже (МСК)."""
    if message.from_user.id not in ADMIN_IDS:
        return

    from wl_traffic.constants import WL_GB_PER_MONTH

    gb = float(WL_GB_PER_MONTH)
    args = (message.text or "").split()
    if len(args) >= 2:
        try:
            gb = float(args[1].strip().replace(",", "."))
        except ValueError:
            await message.answer(
                "❌ Использование: /add_traffic_all [GB]\n"
                "По умолчанию 10 GB. Пример: /add_traffic_all 10"
            )
            return

    if gb <= 0:
        await message.answer("❌ Количество GB должно быть больше 0.")
        return

    user_ids = await sql.add_wl_limit_subscribers_from_today(gb)
    total = len(user_ids)
    if not user_ids:
        await message.answer(
            "Нет пользователей с подпиской до конца сегодня или позже (МСК)."
        )
        return

    await message.answer(
        f"⏳ /add_traffic_all: +{gb:g} GB для {total} пользователей, "
        f"отправка push и проверка squad…"
    )

    traffic_by_username, traffic_by_uuid = await fetch_wl_traffic_gb_for_day(
        x3, retries=1,
    )

    admin_chat_id = message.chat.id
    pushed = 0
    squad_moved = 0
    push_failed = 0
    squad_failed = 0
    skipped_non_tg = 0

    for processed, target_id in enumerate(user_ids, start=1):
        if processed % _ADD_TRAFFIC_ALL_PROGRESS_EVERY == 0:
            try:
                await bot.send_message(
                    admin_chat_id,
                    f"add_traffic_all: {processed} / {total}, "
                    f"push {pushed}, squad → active {squad_moved}",
                )
            except Exception as notify_err:
                logger.warning(
                    "add_traffic_all: не удалось отправить прогресс админу: %s",
                    notify_err,
                )

        trafic_wl, limit_wl = await sql.get_wl_limits(target_id)
        used_gb = await get_wl_used_gb_for_user(
            x3,
            target_id,
            trafic_wl,
            traffic_by_username=traffic_by_username,
            traffic_by_uuid=traffic_by_uuid,
        )
        remaining_gb = max(0.0, round(limit_wl - used_gb, 2))

        panel_user = await fetch_panel_user(x3, target_id)
        if (
            panel_user
            and not user_on_active_squad(panel_user)
            and limit_wl > used_gb
        ):
            if await reassign_to_active_squad(x3, panel_user):
                squad_moved += 1
            else:
                squad_failed += 1

        if target_id <= 0 or not is_telegram_chat_id(target_id):
            skipped_non_tg += 1
            await asyncio.sleep(0.05)
            continue

        try:
            await bot.send_message(
                chat_id=target_id,
                text=lexicon["wl_traffic_admin_grant"].format(
                    gb=gb,
                    limit_gb=limit_wl,
                    used_gb=used_gb,
                    remaining_gb=remaining_gb,
                ),
                parse_mode="HTML",
                reply_markup=create_kb(1, back_to_main=BTN_BACK),
            )
            pushed += 1
        except Exception as e:
            push_failed += 1
            logger.warning(
                "add_traffic_all: push uid=%s: %s",
                target_id,
                e,
            )

        await asyncio.sleep(0.05)

    await message.answer(
        f"✅ <b>Готово (/add_traffic_all)</b>\n\n"
        f"• +{gb:g} GB limit_wl: <b>{total}</b> пользователей\n"
        f"• Push отправлено: <b>{pushed}</b>\n"
        f"• Squad → active (белая нода): <b>{squad_moved}</b>\n"
        f"• Ошибка push: {push_failed}\n"
        f"• Ошибка squad: {squad_failed}\n"
        f"• Пропущено (не Telegram chat_id): {skipped_non_tg}",
        parse_mode="HTML",
    )
    logger.info(
        f"Админ {message.from_user.id}: /add_traffic_all +{gb:g} GB, "
        f"total={total} pushed={pushed} squad_moved={squad_moved}"
    )


@router.message(Command(commands=['reset_bool3']))
async def reset_field_bool_3_all_command(message: Message):
    """Сброс field_bool_3 у всех пользователей (триал / одноразовые акции)."""
    if message.from_user.id not in ADMIN_IDS:
        return
    n = await sql.reset_field_bool_3_all()
    await message.answer(f"Готово: field_bool_3 = false у {n} записей в users.")
    logger.info(f"Админ {message.from_user.id}: сброс field_bool_3 для всех, обновлено строк: {n}")


@router.message(Command(commands=["add_7_sub"]))
async def add_7_sub_command(message: Message):
    """
    Компенсация +7 дней: in_panel=True, is_delete=False, subscription_end_date не пусто.
    Продление в панели (updateClient) и в БД; при успехе — рассылка с кнопкой подключения.
    """
    if message.from_user.id not in ADMIN_IDS:
        return

    user_ids = await sql.select_subscribe_yes()
    total = len(user_ids)
    if not user_ids:
        await message.answer("Нет пользователей: in_panel=True, is_delete=False.")
        return

    await message.answer(
        f"⏳ add_7_sub: обработка {total} пользователей "
        f"(in_panel=True, is_delete=False)…"
    )

    admin_chat_id = message.chat.id
    processed = 0
    extended = 0
    messaged = 0
    skipped_no_sub = 0
    skipped_non_tg = 0
    failed_extend = 0
    failed_message = 0

    for user_id in user_ids:
        processed += 1
        if processed % _ADD7SUB_PROGRESS_EVERY == 0:
            try:
                await bot.send_message(
                    admin_chat_id,
                    f"add_7_sub: обработано {processed} / {total}, "
                    f"продлено {extended}, уведомлено {messaged}",
                )
            except Exception as notify_err:
                logger.warning(
                    "add_7_sub: не удалось отправить прогресс админу: %s",
                    notify_err,
                )

        if not is_telegram_chat_id(user_id):
            skipped_non_tg += 1
            await asyncio.sleep(0.05)
            continue

        user_data = await sql.get_user(user_id)
        if not user_data:
            failed_extend += 1
            await asyncio.sleep(0.1)
            continue

        if user_data[9] is None:
            skipped_no_sub += 1
            await asyncio.sleep(0.05)
            continue

        user_id_str = str(user_id)
        ok = await x3.updateClient(7, user_id_str, user_id)
        if not ok:
            failed_extend += 1
            logger.warning("add_7_sub: не продлили user_id=%s", user_id)
            await asyncio.sleep(0.1)
            continue

        extended += 1
        try:
            await bot.send_message(
                user_id,
                _ADD7SUB_TEXT,
                reply_markup=_ADD7SUB_CONNECT_KB,
            )
            messaged += 1
        except Exception as e:
            failed_message += 1
            logger.warning(
                "add_7_sub: продлено, сообщение не отправлено user_id=%s: %s",
                user_id,
                e,
            )

        await asyncio.sleep(0.1)

    await message.answer(
        "Готово (add_7_sub).\n"
        f"• В выборке: {total}\n"
        f"• Продлено (+7 дн, панель и БД): {extended}\n"
        f"• Уведомлено: {messaged}\n"
        f"• Без subscription_end_date: {skipped_no_sub}\n"
        f"• Ошибка продления: {failed_extend}\n"
        f"• Ошибка сообщения (после продления): {failed_message}\n"
        f"• Пропущено (не Telegram chat_id): {skipped_non_tg}"
    )
    logger.info(
        "Админ %s: add_7_sub total=%s extended=%s messaged=%s",
        message.from_user.id,
        total,
        extended,
        messaged,
    )


@router.message(Command(commands=["add_2_bonus"]))
async def add_2_bonus_command(message: Message):
    """
    Компенсация +2 дня: обычная подписка (subscription_end_date) ≥ 23.06.2026.
    Продление в панели и БД; при успехе — рассылка с кнопкой импорта настроек.
    """
    if message.from_user.id not in ADMIN_IDS:
        return

    user_ids = await sql.select_user_ids_main_sub_on_or_after(_ADD2BONUS_CUTOFF)
    n = len(user_ids)
    if not user_ids:
        await message.answer(
            f"Нет пользователей: обычная подписка заканчивается "
            f"с {_ADD2BONUS_CUTOFF:%d.%m.%Y} и позже."
        )
        return

    confirm_kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="✅ Да, начислить и отправить",
                    callback_data=_ADD2BONUS_YES_CB,
                    style=STYLE_SUCCESS,
                ),
                InlineKeyboardButton(
                    text="❌ Отмена",
                    callback_data=_ADD2BONUS_NO_CB,
                    style=STYLE_DANGER,
                ),
            ]
        ]
    )
    await message.answer(
        f"📋 <b>/add_2_bonus</b>\n\n"
        f"К получателям: <b>{n}</b> чел.\n"
        f"(обычная подписка ≥ {_ADD2BONUS_CUTOFF:%d.%m.%Y}; is_delete=False)\n\n"
        f"Будет начислено <b>+2 дня</b> в панели и БД "
        f"(если подписка истекла — от текущего момента), затем рассылка.\n\n"
        f"Подтвердите отправку.",
        reply_markup=confirm_kb,
    )


@router.callback_query(F.data == _ADD2BONUS_NO_CB)
async def add_2_bonus_cancel(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Нет доступа.", show_alert=True)
        return
    await callback.answer()
    await callback.message.edit_text(
        "Отправка /add_2_bonus отменена.",
        reply_markup=None,
    )


@router.callback_query(F.data == _ADD2BONUS_YES_CB)
async def add_2_bonus_confirm(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Нет доступа.", show_alert=True)
        return

    await callback.answer()
    user_ids = await sql.select_user_ids_main_sub_on_or_after(_ADD2BONUS_CUTOFF)
    if not user_ids:
        await callback.message.edit_text("Список пуст. Повторите /add_2_bonus.")
        return

    total = len(user_ids)
    await callback.message.edit_text(
        f"⏳ /add_2_bonus: обработка {total} пользователей…"
    )

    admin_chat_id = callback.message.chat.id
    processed = 0
    extended = 0
    messaged = 0
    skipped_non_tg = 0
    failed_extend = 0
    failed_message = 0

    for user_id in user_ids:
        processed += 1
        if processed % _ADD2BONUS_PROGRESS_EVERY == 0:
            try:
                await bot.send_message(
                    admin_chat_id,
                    f"add_2_bonus: обработано {processed} / {total}, "
                    f"продлено {extended}, уведомлено {messaged}",
                )
            except Exception as notify_err:
                logger.warning(
                    "add_2_bonus: не удалось отправить прогресс админу: %s",
                    notify_err,
                )

        if not is_telegram_chat_id(user_id):
            skipped_non_tg += 1
            await asyncio.sleep(0.05)
            continue

        user_id_str = str(user_id)
        try:
            ok = await x3.updateClient(2, user_id_str, user_id)
        except Exception as e:
            failed_extend += 1
            logger.warning(
                "add_2_bonus: исключение при продлении user_id=%s: %s",
                user_id,
                e,
            )
            await asyncio.sleep(0.1)
            continue

        if not ok:
            failed_extend += 1
            logger.warning("add_2_bonus: не продлили user_id=%s", user_id)
            await asyncio.sleep(0.1)
            continue

        extended += 1
        try:
            await bot.send_message(
                user_id,
                _ADD2BONUS_TEXT,
                reply_markup=_ADD2BONUS_KB,
            )
            messaged += 1
        except Exception as e:
            failed_message += 1
            logger.warning(
                "add_2_bonus: продлено, сообщение не отправлено user_id=%s: %s",
                user_id,
                e,
            )

        await asyncio.sleep(0.1)

    await callback.message.answer(
        "Готово (/add_2_bonus).\n"
        f"• В выборке: {total}\n"
        f"• Продлено (+2 дн, панель и БД): {extended}\n"
        f"• Уведомлено: {messaged}\n"
        f"• Ошибка продления: {failed_extend}\n"
        f"• Ошибка сообщения (после продления): {failed_message}\n"
        f"• Пропущено (не Telegram chat_id): {skipped_non_tg}"
    )
    logger.info(
        "Админ %s: add_2_bonus total=%s extended=%s messaged=%s",
        callback.from_user.id,
        total,
        extended,
        messaged,
    )


@router.message(Command(commands=['add_7_to_all']))
async def add_7_to_all_command(message: Message):
    """
    Рассылка: нет подписки или она истекла ≥2 дней назад (UTC).
    Кнопка «ТРИАЛ»; +7 дней по нажатию (создание в панели или продление), field_bool_3.
    """
    if message.from_user.id not in ADMIN_IDS:
        return

    user_ids = await sql.SELECT_USER_IDS_NO_ACTIVE_PRO_SUBSCRIPTION()
    n = len(user_ids)
    if not user_ids:
        await message.answer(
            "Нет пользователей: is_delete=False, subscription_end_date пусто "
            "или подписка истекла 2+ дня назад (календарный день UTC)."
        )
        return

    kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="▶️ Превью и подтверждение",
                    callback_data=_ADD7ALL_PREVIEW_CB,
                    style=STYLE_SUCCESS,
                )
            ]
        ]
    )
    await message.answer(
        f"К получателям рассылки: {n} чел.\n"
        f"(is_delete=False; нет подписки или истекла ≥2 дней назад по UTC).\n\n"
        f"Дальше бот пришлёт вам превью текста с кнопкой «🔥Получить ТРИАЛ» и запрос подтверждения.\n"
        f"Начисление +7 дней — только по нажатию: нет в панели → создать на 7 дней, "
        f"есть, но подписка истекла → +7 дней от текущего момента.",
        reply_markup=kb,
    )


@router.callback_query(F.data == _ADD7ALL_PREVIEW_CB)
async def add_7_to_all_preview(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Нет доступа.", show_alert=True)
        return

    await callback.answer()
    user_ids = await sql.SELECT_USER_IDS_NO_ACTIVE_PRO_SUBSCRIPTION()
    n = len(user_ids)
    if not user_ids:
        await callback.message.edit_text("Список пуст. Повторите /add_7_to_all.")
        return

    chat_id = callback.message.chat.id
    await callback.message.edit_text(
        "Ниже — превью рассылки и кнопка подтверждения отправки пользователям."
    )

    await bot.send_message(
        chat_id,
        _ADD7ALL_PROMO_TEXT,
        reply_markup=_ADD7ALL_TRIAL_KB,
    )

    confirm_kb = InlineKeyboardMarkup(
        inline_keyboard=[
            [
                InlineKeyboardButton(
                    text="Да",
                    callback_data=_ADD7ALL_YES_CB,
                    style=STYLE_SUCCESS,
                ),
                InlineKeyboardButton(
                    text="Нет",
                    callback_data=_ADD7ALL_NO_CB,
                    style=STYLE_DANGER,
                ),
            ]
        ]
    )
    await bot.send_message(
        chat_id,
        f"Человек в рассылке — {n}. Подтвердите отправку.",
        reply_markup=confirm_kb,
    )


@router.callback_query(F.data == _ADD7ALL_NO_CB)
async def add_7_to_all_cancel(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Нет доступа.", show_alert=True)
        return
    await callback.answer()
    await callback.message.edit_text(
        "Отправка рассылки add_7_to_all отменена.",
        reply_markup=None,
    )


@router.callback_query(F.data == _ADD7ALL_YES_CB)
async def add_7_to_all_confirm(callback: CallbackQuery):
    if callback.from_user.id not in ADMIN_IDS:
        await callback.answer("Нет доступа.", show_alert=True)
        return

    await callback.answer()
    user_ids = await sql.SELECT_USER_IDS_NO_ACTIVE_PRO_SUBSCRIPTION()
    if not user_ids:
        await callback.message.edit_text("Список пуст. Повторите /add_7_to_all.")
        return

    await callback.message.edit_text(
        f"⏳ Рассылка add_7_to_all: {len(user_ids)} получателей…"
    )

    admin_chat_id = callback.message.chat.id
    sent = 0
    failed = 0
    skipped_non_tg = 0

    for user_id in user_ids:
        if not is_telegram_chat_id(user_id):
            skipped_non_tg += 1
            await asyncio.sleep(0.1)
            continue
        try:
            await bot.send_message(
                user_id,
                _ADD7ALL_PROMO_TEXT,
                reply_markup=_ADD7ALL_TRIAL_KB,
            )
            sent += 1
            if sent % 1000 == 0:
                try:
                    await bot.send_message(
                        admin_chat_id,
                        f"add_7_to_all: отправлено сообщений — {sent}",
                    )
                except Exception as notify_err:
                    logger.warning(
                        "add_7_to_all: не удалось отправить прогресс админу: %s",
                        notify_err,
                    )
        except Exception as e:
            failed += 1
            logger.warning("add_7_to_all: не отправлено user_id=%s: %s", user_id, e)

        await asyncio.sleep(0.1)

    await callback.message.answer(
        "Готово (add_7_to_all).\n"
        f"• Отправлено: {sent}\n"
        f"• Ошибок: {failed}\n"
        f"• Пропущено (не Telegram chat_id): {skipped_non_tg}"
    )


@router.message(Command(commands=['send_push']))
async def send_push_command(message: Message):
    """Отправляет информационное сообщение пользователям, созданным до 16 марта 2026,
    с активной подпиской (in_panel=True, subscription_end_date > now, is_delete=False)."""
    if CHECKER_ID is None or message.from_user.id != CHECKER_ID:
        return

    await message.answer("🔄 Начинаю отправку push-уведомления...")

    # Получаем всех пользователей
    all_users = await sql.get_all_users()

    # Фильтруем
    candidates = [CHECKER_ID]
    for user in all_users:
        if user.is_delete:
            continue
        if not user.in_panel:
            continue
        if user.subscription_end_date:
            continue
        candidates.append(user.user_id)

    if not candidates:
        await message.answer("❌ Нет пользователей, удовлетворяющих условиям.")
        return
    else:
        await message.answer(f"Всего {len(candidates)} пользователей, удовлетворяющих условиям.")

    push_text = '''
🥵 Это была DDoS-атака!

Друзья, простите за временные неудобства. Сервис работает в штатном режиме.
Мы столкнулись с мощной DDoS-атакой, если у вас <b>не открывался личный кабинет — проблема уже решена.</b>

📱 Не можете настроить?
Если вы никак не могли разобраться с импортом конфигов — <b>смотрите видеоинструкцию</b>! Там всё разложено по полочкам.

🌐 Осталось только нажать кнопку "🌐 Подключить Ускоритель соцсетей" — и вы снова в деле.
    '''

    success_count = 0
    fail_count = 0

    for user_id in candidates:
        try:
            user_data = await x3.get_user_by_username(str(user_id))
            if user_data:
                logger.success(f'{user_id} уже в панели')
                raw = user_data['response']
                user = raw[0] if isinstance(raw, list) else raw
                if not isinstance(user, dict):
                    logger.error(f"send_push: неверный формат response для {user_id}")
                    continue

                expire_str = user.get('expireAt')
                if expire_str:
                    try:
                        expire_dt = datetime.fromisoformat(expire_str.replace('Z', '+00:00'))
                        await sql.update_subscription_end_date(user_id, expire_dt)
                    except Exception as e:
                        logger.error(f"send_push: парсинг expireAt для {user_id}: {e}")

                short_uuid = user.get('shortUuid')
                if short_uuid:
                    username_panel = user.get('username') or ''
                    is_white = 'white' in username_panel
                    try:
                        if is_white:
                            await sql.update_white_subscription(user_id, short_uuid)
                        else:
                            await sql.update_subscribtion(user_id, short_uuid)
                    except Exception as e:
                        logger.error(f"send_push: запись shortUuid для {user_id}: {e}")

                continue
            await x3.addClient(5, str(user_id), int(user_id))
            await bot.send_message(user_id,
                                   push_text,
                                   reply_markup=create_kb(
                                       1,
                                       styles={
                                           'video_faq': STYLE_PRIMARY,
                                           'connect_vpn': STYLE_PRIMARY,
                                       },
                                       video_faq='🎥 Видеоинструкция',
                                       connect_vpn='🌐 Подключить Ускоритель соцсетей',
                                   ))
            success_count += 1
            logger.info(f"Push отправлен пользователю {user_id}")
            await asyncio.sleep(0.05)
        except Exception as e:
            fail_count += 1
            logger.error(f"Ошибка отправки для {user_id}: {e}")

    await message.answer(
        f"✅ Рассылка завершена.\n"
        f"👥 Найдено: {len(candidates)}\n"
        f"✅ Успешно: {success_count}\n"
        f"❌ Ошибок: {fail_count}"
    )


_NEW_PANEL_SQUAD_1 = "28b6a3bf-8e81-42dd-9ac8-dab1c9a60b0a"
_NEW_PANEL_SQUAD_2 = "85f8520a-8dd7-40a6-9f27-8a7467096c6a"
_NEW_PANEL_WHITE_SQUAD = "zzz"
_NEW_PANEL_BULK_BATCH = 500


async def _new_panel_bulk_ids(user_ids: list, squad: str) -> tuple[bool, int]:
    """Разбивает id на батчи и вызывает bulk_update_internal_squads."""
    total_affected = 0
    all_ok = True
    for off in range(0, len(user_ids), _NEW_PANEL_BULK_BATCH):
        batch = user_ids[off : off + _NEW_PANEL_BULK_BATCH]
        ok, aff = await x3.bulk_update_internal_squads(batch, [squad])
        total_affected += aff
        if not ok:
            all_ok = False
        await asyncio.sleep(0.15)
    return all_ok, total_affected


@router.message(Command(commands=["new_panel"]))
async def new_panel_command(message: Message):
    """Массовое обновление internal squads: white → white_squad, цифровые username → squad_1/squad_2."""
    if message.from_user.id not in ADMIN_IDS:
        return

    try:
        users = await x3.get_all_panel()
        total_panel = len(users)

        casual_list: list[dict] = []
        white_list: list[dict] = []
        skipped_no_username = 0
        skipped_other = 0

        for u in users:
            un = u.get("username")
            if un is None or str(un).strip() == "":
                skipped_no_username += 1
                continue
            s = str(un)
            if "white" in s:
                white_list.append(u)
            elif s.isdigit():
                casual_list.append(u)
            else:
                skipped_other += 1

        white_ids = [int(u["id"]) for u in white_list if u.get("id") is not None]
        white_no_id = len(white_list) - len(white_ids)
        casual_by_squad: dict[str, list[int]] = {_NEW_PANEL_SQUAD_1: [], _NEW_PANEL_SQUAD_2: []}
        casual_no_id = 0
        for u in casual_list:
            uid = u.get("id")
            if uid is None:
                casual_no_id += 1
                continue
            sq = random.choice([_NEW_PANEL_SQUAD_1, _NEW_PANEL_SQUAD_2])
            casual_by_squad[sq].append(int(uid))

        classified = len(casual_list) + len(white_list)
        bulk_total = len(white_ids) + len(casual_by_squad[_NEW_PANEL_SQUAD_1]) + len(
            casual_by_squad[_NEW_PANEL_SQUAD_2]
        )
        await message.answer(
            f"📋 /new_panel\n"
            f"В панели записей: {total_panel}\n"
            f"По username: обычные — {len(casual_list)}, white — {len(white_list)} "
            f"(всего классифицировано {classified})\n"
            f"К bulk-обновлению (есть id): {bulk_total}\n"
            f"Пропуск: без username — {skipped_no_username}, иной формат username — {skipped_other}\n"
            f"🔄 Начинаю обновление сквадов…"
        )

        white_ok, white_aff = await _new_panel_bulk_ids(white_ids, _NEW_PANEL_WHITE_SQUAD)

        casual_ok = True
        casual_aff = 0
        n_s1 = len(casual_by_squad[_NEW_PANEL_SQUAD_1])
        n_s2 = len(casual_by_squad[_NEW_PANEL_SQUAD_2])
        for sq, ids in casual_by_squad.items():
            if not ids:
                continue
            ok, aff = await _new_panel_bulk_ids(ids, sq)
            casual_aff += aff
            if not ok:
                casual_ok = False

        report = (
            f"✅ /new_panel — отчёт\n"
            f"White: id {len(white_ids)}, affected Σ={white_aff}, "
            f"{'ok' if white_ok else 'были ошибки (см. лог)'}\n"
            f"Casual: squad_1 — {n_s1} юз., squad_2 — {n_s2} юз. "
            f"(random_choice между ними), affected Σ={casual_aff}, "
            f"{'ok' if casual_ok else 'были ошибки (см. лог)'}\n"
        )
        if white_no_id:
            report += f"White без id в панели: {white_no_id}\n"
        if casual_no_id:
            report += f"Casual без id в панели: {casual_no_id}\n"
        await message.answer(report)
        logger.info(
            f"Админ {message.from_user.id} /new_panel: white={len(white_ids)} casual={len(casual_list)} "
            f"white_ok={white_ok} casual_ok={casual_ok}"
        )
    except Exception as e:
        logger.exception("Ошибка в /new_panel")
        await message.answer(f"❌ Ошибка: {str(e)}")
